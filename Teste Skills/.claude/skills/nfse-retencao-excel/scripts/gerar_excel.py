#!/usr/bin/env python3
"""
Gera o relatório final em Excel a partir das notas de NFS-e já classificadas
(JSON do nfse-retencao-classificador).

Uso:
    python gerar_excel.py --input notas_classificadas.json --output relatorio_retencoes.xlsx
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FORMAT = "#,##0.00"

TIPOS_RETENCAO = [
    "IRRF",
    "INSS (Contribuição Previdenciária Retida)",
    "PIS/COFINS/CSLL (Contribuições Sociais Retidas)",
    "ISSQN Retido",
]


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = f"A{row + 1}"


def autosize_columns(ws, min_width=10, max_width=45):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, 0), length)
    for col, width in widths.items():
        ws.column_dimensions[col].width = max(min_width, min(width + 2, max_width))


def build_resumo_sheet(wb, records):
    ws = wb.active
    ws.title = "Resumo"

    headers = [
        "Arquivo", "Número NFS-e", "Situação", "Chave de Acesso", "Competência", "Data Emissão",
        "Prestador", "CNPJ/CPF Prestador", "Tomador", "CNPJ/CPF Tomador",
        "Valor do Serviço", "Valor Líquido", "Tem Retenção",
    ] + TIPOS_RETENCAO + ["Município ISSQN", "Total Retenções"]
    ws.append(headers)

    for record in records:
        retencoes = record.get("retencoes", [])
        retencoes_por_tipo = {r["tipo"]: r["valor"] for r in retencoes}
        municipio_issqn = next((r.get("municipio") for r in retencoes if r["tipo"] == "ISSQN Retido"), None)
        row = [
            record.get("arquivo"),
            record.get("numero_nfse"),
            record.get("situacao"),
            record.get("chave_acesso"),
            record.get("competencia"),
            record.get("data_hora_emissao_nfse"),
            (record.get("prestador") or {}).get("nome"),
            (record.get("prestador") or {}).get("cnpj_cpf"),
            (record.get("tomador") or {}).get("nome"),
            (record.get("tomador") or {}).get("cnpj_cpf"),
            (record.get("valor_total") or {}).get("valor_servico"),
            (record.get("valor_total") or {}).get("valor_liquido"),
            "Sim" if record.get("tem_retencao") else "Não",
        ] + [retencoes_por_tipo.get(tipo) for tipo in TIPOS_RETENCAO] + [municipio_issqn, record.get("valor_total_retencoes")]
        ws.append(row)

    total_col = 15 + len(TIPOS_RETENCAO)  # pula a coluna "Município ISSQN"
    money_cols = [11, 12] + list(range(14, 14 + len(TIPOS_RETENCAO))) + [total_col]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in money_cols:
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FORMAT

    style_header(ws)
    autosize_columns(ws)
    return ws


def build_detalhe_sheet(wb, records):
    ws = wb.create_sheet("Detalhe Retenções")
    headers = [
        "Arquivo", "Número NFS-e", "Situação", "Competência", "Prestador", "CNPJ/CPF Prestador",
        "Tomador", "CNPJ/CPF Tomador", "Tipo de Retenção", "Valor", "Município", "Observação",
    ]
    ws.append(headers)

    for record in records:
        for retencao in record.get("retencoes", []):
            ws.append([
                record.get("arquivo"),
                record.get("numero_nfse"),
                record.get("situacao"),
                record.get("competencia"),
                (record.get("prestador") or {}).get("nome"),
                (record.get("prestador") or {}).get("cnpj_cpf"),
                (record.get("tomador") or {}).get("nome"),
                (record.get("tomador") or {}).get("cnpj_cpf"),
                retencao["tipo"],
                retencao["valor"],
                retencao.get("municipio"),
                retencao["observacao"],
            ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[9]  # coluna "Valor"
        if isinstance(cell.value, (int, float)):
            cell.number_format = MONEY_FORMAT

    style_header(ws)
    autosize_columns(ws)
    return ws


def main():
    parser = argparse.ArgumentParser(description="Gera o Excel final de retenções em NFS-e")
    parser.add_argument("--input", required=True, help="JSON gerado pelo nfse-retencao-classificador")
    parser.add_argument("--output", required=True, help="Caminho do arquivo .xlsx de saída")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERRO: arquivo não encontrado: {input_path}", file=sys.stderr)
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        records = json.load(f)

    wb = Workbook()
    build_resumo_sheet(wb, records)
    build_detalhe_sheet(wb, records)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    com_retencao = sum(1 for r in records if r.get("tem_retencao"))
    print(f"Notas no relatório: {len(records)} ({com_retencao} com retenção)")
    print(f"Excel salvo em: {output_path.resolve()}")


if __name__ == "__main__":
    main()
