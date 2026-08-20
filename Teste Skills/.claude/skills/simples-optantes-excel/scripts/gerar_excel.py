#!/usr/bin/env python3
"""
Gera o relatório final em Excel a partir dos resultados da consulta em
lote de CNPJs na Consulta Optantes (simples-optantes-consulta).

Uso:
    python gerar_excel.py --input resultados.json --output eventos_futuros.xlsx

O JSON de entrada é uma lista de registros no formato retornado pela
skill simples-optantes-consulta, um por CNPJ consultado:

[
  {
    "cnpj": "27882761000170",
    "status": "sucesso",
    "razao_social": "...",
    "situacao_simples": "...",
    "situacao_simei": "...",
    "eventos_futuros": [{"descricao": "...", "data_efeito": "..."}]
  },
  ...
]
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def formatar_cnpj(cnpj: str) -> str:
    if len(cnpj) != 14:
        return cnpj
    return f"{cnpj[0:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = f"A{row + 1}"


def autosize_columns(ws, min_width=10, max_width=50):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, 0), length)
    for col, width in widths.items():
        ws.column_dimensions[col].width = max(min_width, min(width + 2, max_width))


def build_eventos_sheet(wb, records):
    ws = wb.active
    ws.title = "Eventos Futuros"
    headers = [
        "CNPJ", "Razão Social", "Situação Simples Nacional", "Situação SIMEI",
        "Descrição do Evento", "Data Efeito",
    ]
    ws.append(headers)

    for record in records:
        if record.get("status") != "sucesso":
            continue
        for evento in record.get("eventos_futuros") or []:
            ws.append([
                formatar_cnpj(record.get("cnpj", "")),
                record.get("razao_social"),
                record.get("situacao_simples"),
                record.get("situacao_simei"),
                evento.get("descricao"),
                evento.get("data_efeito"),
            ])

    style_header(ws)
    autosize_columns(ws)
    return ws


def build_nao_processados_sheet(wb, records):
    pendentes = [r for r in records if r.get("status") != "sucesso"]
    if not pendentes:
        return None

    ws = wb.create_sheet("Não Localizados")
    ws.append(["CNPJ", "Motivo"])
    for record in pendentes:
        ws.append([formatar_cnpj(record.get("cnpj", "")), record.get("motivo", "não encontrado")])

    style_header(ws)
    autosize_columns(ws)
    return ws


def main():
    parser = argparse.ArgumentParser(description="Gera o Excel de CNPJs com evento futuro no Simples Nacional")
    parser.add_argument("--input", required=True, help="JSON com os resultados da simples-optantes-consulta")
    parser.add_argument("--output", required=True, help="Caminho do arquivo .xlsx de saída")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERRO: arquivo não encontrado: {input_path}", file=sys.stderr)
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        records = json.load(f)

    wb = Workbook()
    build_eventos_sheet(wb, records)
    build_nao_processados_sheet(wb, records)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    com_evento = sum(1 for r in records if r.get("status") == "sucesso" and r.get("eventos_futuros"))
    nao_processados = sum(1 for r in records if r.get("status") != "sucesso")
    total_eventos = sum(len(r.get("eventos_futuros") or []) for r in records if r.get("status") == "sucesso")

    print(f"CNPJs consultados: {len(records)}")
    print(f"CNPJs com evento futuro: {com_evento} ({total_eventos} evento(s) no total)")
    if nao_processados:
        print(f"CNPJs não localizados/com erro: {nao_processados}")
    print(f"Excel salvo em: {output_path.resolve()}")


if __name__ == "__main__":
    main()
